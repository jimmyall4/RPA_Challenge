import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time


def main():
    #turn on and off debug mode
    #debug = True
    debug = False

    if debug == True:
        logging.basicConfig(level=logging.DEBUG)
    else:
        pass
    #open workbook and create wb object
    wb = openpyxl.load_workbook(r"C:\Users\Jimmy\Downloads\challenge.xlsx")

    ws = wb['Sheet1']

    data={}
    url = 'http://www.rpachallenge.com/'

    # open Chrome browser
    logging.debug('Open Browser')
    #update the loaction of chromedriver
    driver = webdriver.Chrome(r'C:\Users\Jimmy\PycharmProjects\RPA_Challenge\venv\Scripts\chromedriver.exe')
    driver.get(url)


    id = 0
    #get the data from excel rows from sheet1
    for row in range(2, 12):

        FirstName = ws['A' + str(row)].value
        LastName = ws['B' + str(row)].value
        CompanyName = ws['C' + str(row)].value
        RoleInCompany = ws['D' + str(row)].value
        Address = ws['E' + str(row)].value
        Email = ws['F' + str(row)].value
        PhoneNumber = ws['G' + str(row)].value

        #add data row to the dictionary
        data[id] ={ 'First':FirstName,'Last':LastName,'Company':CompanyName,'Role':RoleInCompany,'Address':Address, 'Email':Email, 'Phone':PhoneNumber}
        id += 1
    logging.debug('The current id is {}'.format(id))
    #get the number of rows
    numrows = len(data)
    logging.debug('The length of dict is {}'.format(numrows))
    #create counter
    count =0
    logging.debug('The count is {}'.format(count))
    # for i in data:
    #     print(data[i])
    # logging.debug(data.get(count)['Phone'])
#enter in fields on form for each row of data
    clicks = 0

    #start the challenge by clicking start
    try:
        start = driver.find_element_by_xpath("//button[contains(text(),'Start')]")
        start.click()
    except Exception as e:
        logging.debug(e)
    #enter the fields into the proper forms using the dictionary data
    while count < numrows: #continue to repeat the entry of data until there are no more records to key
        try:
            phone = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'Phone')]")
            phone.send_keys(data.get(count)['Phone'])


            lName = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'Last')]")
            lName.send_keys(data.get(count)['Last'])



            fName = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'First')]")
            fName.send_keys(data.get(count)['First'])



            role = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'Role')]")
            role.send_keys(data.get(count)['Role'])



            addy = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'Address')]")
            addy.send_keys(data.get(count)['Address'])

            comp = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'Company')]")
            comp.send_keys(data.get(count)['Company'])

            comp = driver.find_element_by_xpath("//input[contains(@ng-reflect-name,'Email')]")
            comp.send_keys(data.get(count)['Email'])


            logging.debug('All Data Entered, click the submit button')

            #click the submit button
            submit = driver.find_element_by_xpath("//input[@type='submit']")
            submit.click()
            logging.debug('Clicked Submit')

            count+=1
            clicks+=1
            logging.debug(clicks)

        except Exception as e:
            logging.debug(e)

    logging.debug('The total number of clicks was {}'.format(clicks))

    time.sleep(120)

if __name__ == '__main__':
    main()